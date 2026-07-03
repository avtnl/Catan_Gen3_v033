"""
core/strategy_timing.py

Expected-Hand timing for the 142 Catan victory strategies.

Purpose
-------
This module answers the question:

    Given a player's current hand, production, structures, roads, and ports,
    which of the 142 victory strategies appear fastest to reach?

It deliberately does not mutate game, board, or player state. It prepares
strategy resource requirements and delegates turns-to-afford timing to the
shared Expected-Hand engine in core.resource_time_estimator.

Public resource vector convention
---------------------------------
All vectors use the project/game order:

    [Wheat, Ore, Wood, Brick, Sheep]

Main entry points
-----------------
    load_strategy_requirements(...)
    rank_strategies_for_player(...)
    build_strategy_timing_report(...)
    StrategyTimingEngine(...).rank_for_player(...)
    StrategyTimingEngine(...).rank_for_game(...)

Typical usage
-------------
    from core.strategy_timing import StrategyTimingEngine

    engine = StrategyTimingEngine()
    report = engine.rank_for_player(game.board, player, top_n=3)
    print(format_player_top_strategies(report))

Notes for this first draft
--------------------------
- The xlsx/csv file is read once and can be reused by StrategyTimingEngine.
- Existing settlements/cities/roads are treated as progress toward a strategy.
- Road progress is a resource-timing approximation: it counts existing player
  roads, but it does not yet validate whether those roads are on the exact
  legal path to future settlement targets.
- Development-card progress is not subtracted by default because bought/played
  card state is harder to interpret safely from the current Player object.
"""

from __future__ import annotations

from dataclasses import asdict, dataclass, field
from math import isfinite
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple
import csv
import json
import os
import re


# ──────────────────────────────────────────────────────────────────────────────
# Constants and Expected-Hand imports
# ──────────────────────────────────────────────────────────────────────────────
DEFAULT_DEVELOPMENT_CARD_TIMING_MODE: str = "expected"
EXPECTED_DEV_CARD_BUYS_PER_VP_CARD: int = 5
EXPECTED_DEV_CARD_BUYS_PER_KNIGHT: int = 2
BIGGEST_ARMY_KNIGHTS_REQUIRED: int = 3


# Player-trade opportunity reporting.
#
# Report-only:
#   - Does not mutate hands.
#   - Does not mutate need vectors.
#   - Does not change EH timing.
#   - Only adds player_trade_opportunities to top strategy rows.
DEFAULT_INCLUDE_PLAYER_TRADE_OPPORTUNITIES: bool = True

# Trade with another player is only considered if BOTH players have
# victory_points / points <= this threshold.
DEFAULT_MAX_PLAYER_TRADE_VICTORY_POINTS: int = 6

# Counterparty must be expected to have at least this much surplus of
# the resource they would give away.
DEFAULT_MIN_TRADE_SURPLUS_CARDS: float = 1.25

# Keep JSON compact.
DEFAULT_MAX_TRADE_OPPORTUNITIES_PER_STRATEGY: int = 3


# Special-strategy viability gates.
#
# These make Longest Road / Largest Army strategy selection explicit and
# auditable.  A special strategy is not allowed into normal top rankings unless
# the player passes the relevant resource-access and table-competitiveness gates.
DEFAULT_EXCLUDE_NON_VIABLE_SPECIAL_STRATEGIES: bool = True
SPECIAL_CLOSE_TO_BEST_RATIO: float = 0.80
SPECIAL_ACCESS_PIP_THRESHOLD: float = 1.0
SPECIAL_EXCESS_PIPS_THRESHOLD: float = 5.0
SPECIAL_FAVOURABLE_TRADE_RATE: int = 3


try:  # Project constants.
    from core.constants import (  # type: ignore
        NUM_PLAYERS,
        RCARDS_FOR_CITY,
        RCARDS_FOR_DCARD,
        RCARDS_FOR_ROAD,
        RCARDS_FOR_SETTLEMENT,
    )
except Exception:  # Allows import in isolated test environments.
    NUM_PLAYERS = 4
    RCARDS_FOR_CITY = [2, 3, 0, 0, 0]
    RCARDS_FOR_SETTLEMENT = [1, 0, 1, 1, 1]
    RCARDS_FOR_ROAD = [0, 0, 1, 1, 0]
    RCARDS_FOR_DCARD = [1, 1, 0, 0, 1]

try:  # Shared Expected-Hand implementation.
    from core.resource_time_estimator import (  # type: ignore
        EXPECTED_HAND_CONFIDENCE_TARGET,
        EXPECTED_HAND_CONTINUOUS_TRADING,
        EXPECTED_HAND_MAX_TURNS,
        EXPECTED_HAND_REQUIRE_CONFIDENCE,
        EXPECTED_HAND_ROLLS_PER_PLAYER_TURN,
        EXPECTED_HAND_STEP,
        INFINITE_TURNS,
        RESOURCE_NAMES,
        clean_int_vector,
        clean_vector,
        compute_payability_with_trades,
        estimate_expected_hand_after_turns,
        estimate_first_payable_turn,
        finite_or_9999,
        get_player_ports,
        get_player_production_pips,
        get_player_resource_cards_vector,
        get_player_trade_rates,
        safe_float,
        safe_int,
        vector_sum,
        vector_to_named_dict,
    )
    _RTE_AVAILABLE = True
except Exception:  # pragma: no cover - only used outside the project.
    _RTE_AVAILABLE = False
    EXPECTED_HAND_CONFIDENCE_TARGET = 0.85
    EXPECTED_HAND_CONTINUOUS_TRADING = True
    EXPECTED_HAND_MAX_TURNS = 60.0
    EXPECTED_HAND_REQUIRE_CONFIDENCE = False
    EXPECTED_HAND_ROLLS_PER_PLAYER_TURN = NUM_PLAYERS
    EXPECTED_HAND_STEP = 0.25
    INFINITE_TURNS = 9999.0
    RESOURCE_NAMES = ["Wheat", "Ore", "Wood", "Brick", "Sheep"]

    def safe_float(value: Any, default: float = 0.0) -> float:
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    def safe_int(value: Any, default: int = 0) -> int:
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return default

    def clean_vector(values: Optional[Sequence[Any]], length: int = 5, default: float = 0.0) -> List[float]:
        if values is None:
            return [default] * length
        out = [safe_float(v, default) for v in list(values)[:length]]
        if len(out) < length:
            out.extend([default] * (length - len(out)))
        return out

    def clean_int_vector(values: Optional[Sequence[Any]], length: int = 5, default: int = 4) -> List[int]:
        out = [max(1, safe_int(v, default)) for v in list(values or [])[:length]]
        if len(out) < length:
            out.extend([default] * (length - len(out)))
        return out

    def vector_sum(values: Sequence[Any]) -> float:
        return float(sum(clean_vector(values)))

    def finite_or_9999(value: Any) -> float:
        v = safe_float(value, INFINITE_TURNS)
        return v if isfinite(v) else INFINITE_TURNS

    def vector_to_named_dict(values: Sequence[Any], *, digits: Optional[int] = None) -> Dict[str, float]:
        vec = clean_vector(values)
        if digits is not None:
            vec = [round(v, digits) for v in vec]
        return {RESOURCE_NAMES[i]: vec[i] for i in range(5)}

    def _missing_rte(*_: Any, **__: Any) -> Any:
        raise ImportError(
            "core.resource_time_estimator is required for strategy timing. "
            "Place this file inside your project's core/ folder."
        )

    estimate_first_payable_turn = _missing_rte
    estimate_expected_hand_after_turns = _missing_rte
    compute_payability_with_trades = _missing_rte
    get_player_ports = lambda board, player: []  # noqa: E731
    get_player_production_pips = _missing_rte
    get_player_resource_cards_vector = _missing_rte
    get_player_trade_rates = _missing_rte


RESOURCE_ORDER_NAMES: Tuple[str, str, str, str, str] = ("Wheat", "Ore", "Wood", "Brick", "Sheep")
DEFAULT_STRATEGY_FILENAMES: Tuple[str, str] = (
    "catan_142_ways_resource_requirements.xlsx",
    "catan_142_ways_resource_requirements.csv",
)


# ──────────────────────────────────────────────────────────────────────────────
# Small utilities
# ──────────────────────────────────────────────────────────────────────────────

def _as_bool(value: Any) -> bool:
    text = str(value or "").strip().lower()
    return text in {"1", "true", "t", "yes", "y", "ja", "x"}


def _norm_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _norm_key_alias(key: str) -> str:
    aliases = {
        "way": "way_id",
        "id": "way_id",
        "wayid": "way_id",
        "source": "source_row",
        "longestroad": "longest_road",
        "longest_road_flag": "longest_road",
        "biggestarmy": "biggest_army",
        "largest_army": "biggest_army",
        "victory_cards": "victory_point_cards",
        "vp_cards": "victory_point_cards",
        "vp_card": "victory_point_cards",
        "total_vp": "total_victory_points",
        "twelve_point": "twelve_point_edge_case",
        "12_point_edge_case": "twelve_point_edge_case",
        "article_cost": "article_min_cost",
        "new_settlements": "new_settlements_to_build",
        "new_settlement_to_build": "new_settlements_to_build",
        "city_upgrade": "city_upgrades",
        "roads": "roads_to_build",
        "road": "roads_to_build",
        "development_cards": "development_cards_to_buy",
        "dev_cards": "development_cards_to_buy",
        "dcards": "development_cards_to_buy",
        "wheat": "wheat_needed",
        "ore": "ore_needed",
        "wood": "wood_needed",
        "brick": "brick_needed",
        "wool": "wool_needed",
        "sheep_needed": "wool_needed",
        "sheep": "wool_needed",
    }
    return aliases.get(key, key)


def _row_get(row: Mapping[str, Any], *keys: str, default: Any = 0) -> Any:
    for key in keys:
        normalized = _norm_key_alias(_norm_header(key))
        if normalized in row:
            value = row[normalized]
            if value is not None and str(value).strip() != "":
                return value
    return default


def _tuple5(values: Sequence[Any]) -> Tuple[float, float, float, float, float]:
    vec = clean_vector(values)
    return (float(vec[0]), float(vec[1]), float(vec[2]), float(vec[3]), float(vec[4]))


def _int_tuple5(values: Sequence[Any]) -> Tuple[int, int, int, int, int]:
    vec = [int(round(x)) for x in clean_vector(values)]
    return (vec[0], vec[1], vec[2], vec[3], vec[4])


def _vector_mul(cost: Sequence[Any], count: int) -> List[float]:
    c = clean_vector(cost)
    return [float(count) * c[i] for i in range(5)]


def _vector_add_many(vectors: Iterable[Sequence[Any]]) -> List[float]:
    total = [0.0] * 5
    for vec in vectors:
        cv = clean_vector(vec)
        for idx in range(5):
            total[idx] += cv[idx]
    return total


def _resource_dict(values: Sequence[Any], *, digits: Optional[int] = None) -> Dict[str, float]:
    # Use canonical names even if RESOURCE_NAMES ever differs in an older module.
    vec = clean_vector(values)
    if digits is not None:
        vec = [round(v, digits) for v in vec]
    return {RESOURCE_ORDER_NAMES[i]: vec[i] for i in range(5)}


def _display_need(values: Sequence[Any]) -> str:
    vec = clean_vector(values)
    parts = []
    labels = ["W", "O", "Wd", "B", "S"]
    for idx, value in enumerate(vec):
        if abs(value) > 1e-9:
            as_int = int(round(value))
            shown = as_int if abs(value - as_int) < 1e-9 else round(value, 2)
            parts.append(f"{labels[idx]}{shown}")
    return " ".join(parts) if parts else "none"


def _safe_sorted_ids(values: Iterable[Any]) -> List[int]:
    out = []
    for value in values or []:
        try:
            out.append(int(value))
        except Exception:
            continue
    return sorted(set(out))


def _canonical_road_id(road: Any) -> Optional[Tuple[int, int]]:
    try:
        a, b = tuple(road)
        ai = int(a)
        bi = int(b)
    except Exception:
        return None
    if ai == bi:
        return None
    return (min(ai, bi), max(ai, bi))


def _player_road_count(player: Any) -> int:
    roads = set()
    for road in getattr(player, "roads", []) or []:
        rid = _canonical_road_id(road)
        if rid is not None:
            roads.add(rid)
    return len(roads)


def _player_dev_card_progress(player: Any) -> int:
    """Best-effort count of bought/held dev cards. Off by default in timing."""
    values = [0]

    try:
        values.append(len(getattr(player, "development_cards", []) or []))
    except Exception:
        pass

    values.append(safe_int(getattr(player, "number_of_dcards", 0), 0))

    # dcard_summary shape in this project is a list of rows such as
    # [name, bought/held/played counts...]. The semantics can vary between
    # versions, so we only use a very conservative row sum fallback.
    try:
        summary = getattr(player, "dcard_summary", []) or []
        conservative = 0
        for row in summary:
            if isinstance(row, (list, tuple)) and len(row) >= 2:
                conservative += max(0, safe_int(row[1], 0))
        values.append(conservative)
    except Exception:
        pass

    return max(values)


def strategy_cost_from_components(
    *,
    new_settlements: int = 0,
    city_upgrades: int = 0,
    roads: int = 0,
    dev_cards: int = 0,
) -> Tuple[float, float, float, float, float]:
    """Return total resource vector from build counts in game order."""
    total = _vector_add_many(
        [
            _vector_mul(RCARDS_FOR_SETTLEMENT, max(0, int(new_settlements))),
            _vector_mul(RCARDS_FOR_CITY, max(0, int(city_upgrades))),
            _vector_mul(RCARDS_FOR_ROAD, max(0, int(roads))),
            _vector_mul(RCARDS_FOR_DCARD, max(0, int(dev_cards))),
        ]
    )
    return _tuple5(total)


def expected_development_card_buys(
    *,
    victory_point_cards: int = 0,
    largest_army: bool = False,
    listed_development_cards: int = 0,
    mode: str = DEFAULT_DEVELOPMENT_CARD_TIMING_MODE,
) -> int:
    """
    Convert desired development-card outcomes into expected card purchases.

    In the strategy table, a row may say it needs 1 Victory Point card or
    Largest Army. Those are outcomes, not guaranteed purchases. This function
    estimates the number of development cards that must be bought on average:

        1 Victory Point card -> 5 development-card buys
        Largest Army -> 3 knights -> 2 buys per knight -> 6 buys

    The spreadsheet's listed Development_Cards_To_Buy is kept as a lower bound
    so older strategy rows remain compatible.
    """
    listed = max(0, int(listed_development_cards))
    if str(mode or "").strip().lower() in {"listed", "raw", "minimum"}:
        return listed

    vp_buys = max(0, int(victory_point_cards)) * int(EXPECTED_DEV_CARD_BUYS_PER_VP_CARD)
    army_buys = (
        int(BIGGEST_ARMY_KNIGHTS_REQUIRED) * int(EXPECTED_DEV_CARD_BUYS_PER_KNIGHT)
        if bool(largest_army)
        else 0
    )
    return max(listed, vp_buys + army_buys)


# ──────────────────────────────────────────────────────────────────────────────
# Data containers
# ──────────────────────────────────────────────────────────────────────────────

@dataclass(frozen=True)
class StrategyRequirement:
    way_id: int
    source_row: int = 0
    longest_road: bool = False
    biggest_army: bool = False
    cities: int = 0
    settlements: int = 0
    victory_point_cards: int = 0
    total_victory_points: int = 0
    twelve_point_edge_case: bool = False
    article_min_cost: int = 0
    buildings: int = 0
    new_settlements_to_build: int = 0
    city_upgrades: int = 0
    roads_to_build: int = 0
    development_cards_to_buy: int = 0
    static_need: Tuple[float, float, float, float, float] = (0.0, 0.0, 0.0, 0.0, 0.0)
    calculated_need: Tuple[float, float, float, float, float] = (0.0, 0.0, 0.0, 0.0, 0.0)
    production_only_total: float = 0.0
    production_minus_article_min: float = 0.0
    validation_warnings: Tuple[str, ...] = field(default_factory=tuple)
    raw: Mapping[str, Any] = field(default_factory=dict)

    @property
    def target_total_buildings(self) -> int:
        if self.buildings > 0:
            return int(self.buildings)
        return int(self.cities + self.settlements)

    @property
    def tags(self) -> List[str]:
        tags: List[str] = []
        if self.longest_road:
            tags.append("Longest Road")
        if self.biggest_army:
            tags.append("Largest Army")
        if self.cities:
            tags.append(f"{self.cities} cities")
        if self.settlements:
            tags.append(f"{self.settlements} settlements")
        if self.victory_point_cards:
            tags.append(f"{self.victory_point_cards} VP cards")
        if self.development_cards_to_buy:
            tags.append(f"{self.development_cards_to_buy} dev cards")
        if self.twelve_point_edge_case:
            tags.append("12-point edge case")
        return tags

    def as_dict(self) -> Dict[str, Any]:
        data = asdict(self)
        # Public report naming: Catan uses "Largest Army".
        # Keep the internal field for backward-compatible spreadsheet loading.
        data["largest_army"] = data.pop("biggest_army", False)
        data["static_need_named"] = _resource_dict(self.static_need)
        data["calculated_need_named"] = _resource_dict(self.calculated_need)
        data["tags"] = self.tags
        return data


@dataclass(frozen=True)
class PlayerStrategyState:
    player_id: int
    color: str
    current_hand: Tuple[float, float, float, float, float]
    production_pips: Tuple[float, float, float, float, float]
    trade_rates: Tuple[int, int, int, int, int]
    ports: Tuple[str, ...]
    settlements: Tuple[int, ...]
    cities: Tuple[int, ...]
    roads_count: int
    dev_card_progress: int = 0

    def as_dict(self) -> Dict[str, Any]:
        return {
            "player_id": self.player_id,
            "color": self.color,
            "current_hand": self.current_hand,
            "current_hand_named": _resource_dict(self.current_hand),
            "production_pips": self.production_pips,
            "production_pips_named": _resource_dict(self.production_pips, digits=2),
            "trade_rates": self.trade_rates,
            "trade_rates_named": {RESOURCE_ORDER_NAMES[i]: self.trade_rates[i] for i in range(5)},
            "ports": list(self.ports),
            "settlements": list(self.settlements),
            "cities": list(self.cities),
            "roads_count": self.roads_count,
            "dev_card_progress": self.dev_card_progress,
        }


@dataclass(frozen=True)
class SpecialStrategyViability:
    """Transparent pass/fail result for Longest Road / Largest Army gates."""

    strategy_name: str
    viable: bool
    status: str
    reason: str
    core_pips: Mapping[str, float]
    core_total: float
    best_table_core_total: float
    close_to_best_ratio: float
    close_to_best_required: float
    close_to_best: bool
    missing_resources: Tuple[str, ...]
    favourable_trade_support: bool
    trade_support_resources: Tuple[str, ...]
    passed_rules: Tuple[str, ...]
    failed_rules: Tuple[str, ...]

    def as_dict(self) -> Dict[str, Any]:
        return {
            "strategy_name": self.strategy_name,
            "viable": self.viable,
            "status": self.status,
            "reason": self.reason,
            "core_pips": dict(self.core_pips),
            "core_total": self.core_total,
            "best_table_core_total": self.best_table_core_total,
            "close_to_best_ratio": self.close_to_best_ratio,
            "close_to_best_required": self.close_to_best_required,
            "close_to_best": self.close_to_best,
            "missing_resources": list(self.missing_resources),
            "favourable_trade_support": self.favourable_trade_support,
            "trade_support_resources": list(self.trade_support_resources),
            "passed_rules": list(self.passed_rules),
            "failed_rules": list(self.failed_rules),
        }


@dataclass(frozen=True)
class StrategyRemainingNeed:
    way_id: int
    remaining_new_settlements: int
    remaining_city_upgrades: int
    remaining_roads_to_build: int
    remaining_dev_cards_to_buy: int
    need_vector: Tuple[float, float, float, float, float]
    total_cards: float
    progress: Mapping[str, Any] = field(default_factory=dict)
    warnings: Tuple[str, ...] = field(default_factory=tuple)

    def as_dict(self) -> Dict[str, Any]:
        return {
            "way_id": self.way_id,
            "remaining_new_settlements": self.remaining_new_settlements,
            "remaining_city_upgrades": self.remaining_city_upgrades,
            "remaining_roads_to_build": self.remaining_roads_to_build,
            "remaining_dev_cards_to_buy": self.remaining_dev_cards_to_buy,
            "need_vector": self.need_vector,
            "need_named": _resource_dict(self.need_vector),
            "need_compact": _display_need(self.need_vector),
            "total_cards": self.total_cards,
            "progress": dict(self.progress),
            "warnings": list(self.warnings),
        }


@dataclass(frozen=True)
class StrategyTimingRow:
    way_id: int
    rank: int
    turns: float
    found: bool
    confidence: float
    confidence_label: str
    need_vector: Tuple[float, float, float, float, float]
    expected_hand: Tuple[float, float, float, float, float]
    trade_rates: Tuple[int, int, int, int, int]
    tags: Tuple[str, ...]
    reason: str
    bottlenecks: Tuple[str, ...]
    trade_dependency: Tuple[str, ...]
    strategy: StrategyRequirement
    remaining: StrategyRemainingNeed
    estimate: Mapping[str, Any] = field(default_factory=dict)
    special_viability: Mapping[str, Any] = field(default_factory=dict)
    trade_diagnostics: Mapping[str, Any] = field(default_factory=dict)

    def as_dict(self, *, include_debug: bool = False) -> Dict[str, Any]:
        out = {
            "rank": self.rank,
            "way_id": self.way_id,
            "turns": self.turns,
            "found": self.found,
            "confidence": self.confidence,
            "confidence_label": self.confidence_label,
            "need_vector": self.need_vector,
            "need_named": _resource_dict(self.need_vector),
            "need_compact": _display_need(self.need_vector),
            "expected_hand": self.expected_hand,
            "expected_hand_named": _resource_dict(self.expected_hand, digits=2),
            "trade_rates": self.trade_rates,
            "tags": list(self.tags),
            "reason": self.reason,
            "bottlenecks": list(self.bottlenecks),
            "trade_dependency": list(self.trade_dependency),
            "trade_diagnostics": dict(self.trade_diagnostics),
            "special_viability": dict(self.special_viability),
            "remaining": self.remaining.as_dict(),
            "strategy_summary": {
                "longest_road": self.strategy.longest_road,
                "largest_army": self.strategy.biggest_army,
                "cities": self.strategy.cities,
                "settlements": self.strategy.settlements,
                "victory_point_cards": self.strategy.victory_point_cards,
                "development_cards_to_buy": self.strategy.development_cards_to_buy,
                "article_min_cost": self.strategy.article_min_cost,
                "total_victory_points": self.strategy.total_victory_points,
                "twelve_point_edge_case": self.strategy.twelve_point_edge_case,
            },
        }
        if include_debug:
            out["strategy"] = self.strategy.as_dict()
            out["estimate"] = dict(self.estimate)
        return out


# ──────────────────────────────────────────────────────────────────────────────
# Loader
# ──────────────────────────────────────────────────────────────────────────────

def _read_csv_rows(path: Path) -> List[Dict[str, Any]]:
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = []
        for row in reader:
            norm = {_norm_key_alias(_norm_header(k)): v for k, v in row.items()}
            rows.append(norm)
        return rows


def _read_xlsx_rows(path: Path, *, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "Reading .xlsx strategy requirements requires openpyxl. "
            "Install openpyxl or export the workbook as catan_142_ways_resource_requirements.csv."
        ) from exc

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

        raw_rows = list(ws.iter_rows(values_only=True))
        header_index = None
        headers: List[str] = []

        for idx, raw in enumerate(raw_rows):
            candidate = [_norm_key_alias(_norm_header(cell)) for cell in raw]
            if "way_id" in candidate and any(h.endswith("needed") for h in candidate):
                header_index = idx
                headers = candidate
                break

        if header_index is None:
            raise ValueError(
                f"Could not find a header row in {path.name}. Expected columns such as Way_ID and Wheat_Needed."
            )

        rows: List[Dict[str, Any]] = []
        for raw in raw_rows[header_index + 1 :]:
            if not raw or all(value is None or str(value).strip() == "" for value in raw):
                continue
            row = {}
            for col_idx, key in enumerate(headers):
                if not key:
                    continue
                value = raw[col_idx] if col_idx < len(raw) else None
                row[key] = value
            rows.append(row)

        return rows
    finally:
        wb.close()


def _row_to_requirement(row: Mapping[str, Any], fallback_source_row: int) -> StrategyRequirement:
    way_id = safe_int(_row_get(row, "way_id", default=fallback_source_row), fallback_source_row)

    static_need = _tuple5(
        [
            _row_get(row, "wheat_needed", default=0),
            _row_get(row, "ore_needed", default=0),
            _row_get(row, "wood_needed", default=0),
            _row_get(row, "brick_needed", default=0),
            _row_get(row, "wool_needed", "sheep_needed", default=0),
        ]
    )

    new_settlements = safe_int(_row_get(row, "new_settlements_to_build", default=0), 0)
    city_upgrades = safe_int(_row_get(row, "city_upgrades", default=0), 0)
    roads = safe_int(_row_get(row, "roads_to_build", default=0), 0)
    largest_army_tf = _as_bool(_row_get(row, "largest_army", "biggest_army", default=False))
    vp_cards = safe_int(_row_get(row, "victory_point_cards", default=0), 0)
    listed_dev_cards = safe_int(_row_get(row, "development_cards_to_buy", default=0), 0)
    dev_cards = expected_development_card_buys(
        victory_point_cards=vp_cards,
        largest_army=largest_army_tf,
        listed_development_cards=listed_dev_cards,
    )

    calculated = strategy_cost_from_components(
        new_settlements=new_settlements,
        city_upgrades=city_upgrades,
        roads=roads,
        dev_cards=dev_cards,
    )

    warnings: List[str] = []
    if dev_cards != listed_dev_cards:
        warnings.append(
            "Development-card timing adjusted for probability: "
            f"listed={listed_dev_cards}, expected_buys={dev_cards}"
        )
    if _int_tuple5(static_need) != _int_tuple5(calculated):
        warnings.append(
            "Spreadsheet resource columns differ from component-calculated cost: "
            f"spreadsheet={_int_tuple5(static_need)}, calculated={_int_tuple5(calculated)}"
        )

    article_min_cost = safe_int(_row_get(row, "article_min_cost", default=0), 0)
    production_only_total = safe_float(_row_get(row, "production_only_total", default=vector_sum(static_need)), vector_sum(static_need))
    if article_min_cost and abs(production_only_total - vector_sum(static_need)) > 1e-9:
        warnings.append(
            "Production_Only_Total differs from sum of resource columns: "
            f"production_only_total={production_only_total}, sum={vector_sum(static_need)}"
        )

    return StrategyRequirement(
        way_id=way_id,
        source_row=safe_int(_row_get(row, "source_row", default=fallback_source_row), fallback_source_row),
        longest_road=_as_bool(_row_get(row, "longest_road", default=False)),
        biggest_army=largest_army_tf,
        cities=safe_int(_row_get(row, "cities", default=0), 0),
        settlements=safe_int(_row_get(row, "settlements", default=0), 0),
        victory_point_cards=vp_cards,
        total_victory_points=safe_int(_row_get(row, "total_victory_points", default=0), 0),
        twelve_point_edge_case=_as_bool(_row_get(row, "twelve_point_edge_case", default=False)),
        article_min_cost=article_min_cost,
        buildings=safe_int(_row_get(row, "buildings", default=0), 0),
        new_settlements_to_build=new_settlements,
        city_upgrades=city_upgrades,
        roads_to_build=roads,
        development_cards_to_buy=dev_cards,
        static_need=static_need,
        calculated_need=calculated,
        production_only_total=production_only_total,
        production_minus_article_min=safe_float(_row_get(row, "production_minus_article_min", default=0), 0.0),
        validation_warnings=tuple(warnings),
        raw=dict(row),
    )


def resolve_strategy_requirements_path(
    path: Optional[os.PathLike[str] | str] = None,
    *,
    base_dir: Optional[os.PathLike[str] | str] = None,
) -> Path:
    """
    Resolve the 142-way requirements file.

    Search order when path is None:
      1. base_dir, if provided
      2. project root inferred from this file: core/strategy_timing.py -> parent.parent
      3. current working directory

    Preferred filename is .xlsx, with .csv as fallback.
    """
    if path is not None:
        candidate = Path(path).expanduser()
        if candidate.is_file():
            return candidate.resolve()
        raise FileNotFoundError(f"Strategy requirements file not found: {candidate}")

    search_dirs: List[Path] = []
    if base_dir is not None:
        search_dirs.append(Path(base_dir).expanduser())

    try:
        # This module is expected to live at project_root/core/strategy_timing.py.
        search_dirs.append(Path(__file__).resolve().parent.parent)
    except Exception:
        pass

    search_dirs.append(Path.cwd())

    seen = set()
    unique_dirs = []
    for directory in search_dirs:
        try:
            resolved = directory.resolve()
        except Exception:
            resolved = directory
        if resolved in seen:
            continue
        seen.add(resolved)
        unique_dirs.append(resolved)

    for directory in unique_dirs:
        for filename in DEFAULT_STRATEGY_FILENAMES:
            candidate = directory / filename
            if candidate.is_file():
                return candidate.resolve()

    searched = ", ".join(str(d) for d in unique_dirs)
    raise FileNotFoundError(
        "Could not find catan_142_ways_resource_requirements.xlsx or .csv. "
        f"Searched: {searched}"
    )


def load_strategy_requirements(
    path: Optional[os.PathLike[str] | str] = None,
    *,
    base_dir: Optional[os.PathLike[str] | str] = None,
    sheet_name: Optional[str] = None,
    sort_by_way_id: bool = True,
) -> List[StrategyRequirement]:
    """Load and normalize the 142 strategy requirements from xlsx or csv."""
    resolved = resolve_strategy_requirements_path(path, base_dir=base_dir)

    suffix = resolved.suffix.lower()
    if suffix == ".csv":
        rows = _read_csv_rows(resolved)
    elif suffix in {".xlsx", ".xlsm"}:
        rows = _read_xlsx_rows(resolved, sheet_name=sheet_name)
    else:
        raise ValueError(f"Unsupported strategy requirements file type: {resolved.suffix}")

    requirements = [_row_to_requirement(row, idx) for idx, row in enumerate(rows, start=1)]

    if not requirements:
        raise ValueError(f"No strategy requirements were loaded from {resolved}")

    if sort_by_way_id:
        requirements.sort(key=lambda s: s.way_id)

    return requirements


# ──────────────────────────────────────────────────────────────────────────────
# Player state and dynamic remaining need
# ──────────────────────────────────────────────────────────────────────────────

def build_player_strategy_state(board: Any, player: Any) -> PlayerStrategyState:
    """Extract the player state needed by the strategy timing engine."""
    hand = _tuple5(get_player_resource_cards_vector(player))
    pips = _tuple5(get_player_production_pips(board, player))
    rates = tuple(clean_int_vector(get_player_trade_rates(board, player), default=4))

    try:
        ports = tuple(sorted(str(p) for p in get_player_ports(board, player)))
    except Exception:
        ports = tuple()

    cities = tuple(_safe_sorted_ids(getattr(player, "cities", []) or []))
    settlements_raw = _safe_sorted_ids(getattr(player, "settlements", []) or [])
    city_set = set(cities)
    settlements = tuple(i for i in settlements_raw if i not in city_set)

    return PlayerStrategyState(
        player_id=safe_int(getattr(player, "id", 0), 0),
        color=str(getattr(player, "color", "")),
        current_hand=hand,
        production_pips=pips,
        trade_rates=rates,  # type: ignore[arg-type]
        ports=ports,
        settlements=settlements,
        cities=cities,
        roads_count=_player_road_count(player),
        dev_card_progress=_player_dev_card_progress(player),
    )


# ──────────────────────────────────────────────────────────────────────────────
# Special-strategy viability gates
# ──────────────────────────────────────────────────────────────────────────────

def _resource_index(name: str) -> int:
    lookup = {resource.lower(): idx for idx, resource in enumerate(RESOURCE_ORDER_NAMES)}
    # Backward-compatible aliases.
    lookup.setdefault("wool", lookup.get("sheep", 4))
    lookup.setdefault("sheep", lookup.get("sheep", 4))
    return lookup[str(name).strip().lower()]


def _has_meaningful_access(pips: Any) -> bool:
    return safe_float(pips, 0.0) >= SPECIAL_ACCESS_PIP_THRESHOLD


def _status_from_viability(*, viable: bool, direct_all: bool, trade_ok: bool) -> str:
    if not viable:
        return "not_viable"
    if direct_all:
        return "strong"
    if trade_ok:
        return "possible_with_trade"
    return "weak"


def _find_favourable_trade_support(
    player_state: PlayerStrategyState,
    *,
    core_resource_names: Iterable[str],
) -> Tuple[bool, Tuple[str, ...]]:
    """
    Return whether the player has favourable trade support.

    Favourable trade support is intentionally stricter than ordinary 4:1 bank
    trading. It requires:
        - a rate better than 4:1, currently <= 3:1; and
        - meaningful excess production in the export resource.

    This supports human-like strategy selection: a player should not choose
    Largest Army or Longest Road just because a distant 4:1 bank trade is
    mathematically possible.
    """
    support: List[str] = []
    pips = clean_vector(player_state.production_pips)
    rates = clean_int_vector(player_state.trade_rates, default=4)

    for idx, name in enumerate(RESOURCE_ORDER_NAMES):
        pip_value = safe_float(pips[idx], 0.0)
        rate = safe_int(rates[idx], 4)

        if rate <= SPECIAL_FAVOURABLE_TRADE_RATE and pip_value >= SPECIAL_EXCESS_PIPS_THRESHOLD:
            support.append(f"{name} at {rate}:1")

    return bool(support), tuple(support)


def _best_core_total(
    all_player_states: Sequence[PlayerStrategyState],
    resource_indices: Sequence[int],
) -> float:
    if not all_player_states:
        return 0.0
    return max(
        sum(safe_float(state.production_pips[idx], 0.0) for idx in resource_indices)
        for state in all_player_states
    )


def _evaluate_core_special_viability(
    *,
    strategy_name: str,
    player_state: PlayerStrategyState,
    all_player_states: Sequence[PlayerStrategyState],
    core_resource_names: Sequence[str],
    direct_all_rule: str,
    partial_with_trade_rule: str,
    insufficient_access_rule: str,
    close_to_best_rule: str,
    not_close_to_best_rule: str,
) -> SpecialStrategyViability:
    indices = [_resource_index(name) for name in core_resource_names]
    pips_by_name = {
        name: safe_float(player_state.production_pips[idx], 0.0)
        for name, idx in zip(core_resource_names, indices)
    }
    core_total = round(sum(pips_by_name.values()), 3)

    best_total = _best_core_total(all_player_states, indices)
    if best_total <= 1e-9:
        # No player has the required resource profile. Keep the gate strict.
        close_required = SPECIAL_ACCESS_PIP_THRESHOLD * len(indices)
    else:
        close_required = best_total * SPECIAL_CLOSE_TO_BEST_RATIO
    close_required = round(close_required, 3)
    close_to_best = core_total >= close_required

    missing = tuple(
        name for name, value in pips_by_name.items()
        if not _has_meaningful_access(value)
    )

    direct_all = len(missing) == 0
    partial_allowed = len(missing) <= 1
    trade_ok, trade_support = _find_favourable_trade_support(
        player_state,
        core_resource_names=core_resource_names,
    )

    passed: List[str] = []
    failed: List[str] = []

    if direct_all:
        passed.append(direct_all_rule)
    elif partial_allowed and trade_ok:
        passed.append(partial_with_trade_rule)
    else:
        failed.append(insufficient_access_rule)

    if close_to_best:
        passed.append(close_to_best_rule)
    else:
        failed.append(not_close_to_best_rule)

    viable = len(failed) == 0
    status = _status_from_viability(viable=viable, direct_all=direct_all, trade_ok=trade_ok)

    if viable and direct_all:
        reason = (
            f"{strategy_name} viable: "
            f"core_pips={pips_by_name}, core_total={core_total}, "
            f"close-to-best threshold={close_required}, and all core resources are available."
        )
    elif viable:
        reason = (
            f"{strategy_name} possible with favourable trade: "
            f"core_pips={pips_by_name}, core_total={core_total}, "
            f"close-to-best threshold={close_required}, missing={list(missing)}, "
            f"trade_support={list(trade_support)}."
        )
    else:
        reason = (
            f"{strategy_name} rejected: "
            f"core_pips={pips_by_name}, core_total={core_total}, "
            f"required close-to-best={close_required}, missing={list(missing)}, "
            f"trade_support={list(trade_support)}, failed_rules={failed}."
        )

    return SpecialStrategyViability(
        strategy_name=strategy_name,
        viable=viable,
        status=status,
        reason=reason,
        core_pips=pips_by_name,
        core_total=core_total,
        best_table_core_total=round(best_total, 3),
        close_to_best_ratio=SPECIAL_CLOSE_TO_BEST_RATIO,
        close_to_best_required=close_required,
        close_to_best=close_to_best,
        missing_resources=missing,
        favourable_trade_support=trade_ok,
        trade_support_resources=trade_support,
        passed_rules=tuple(passed),
        failed_rules=tuple(failed),
    )


def evaluate_longest_road_viability(
    player_state: PlayerStrategyState,
    all_player_states: Sequence[PlayerStrategyState],
) -> SpecialStrategyViability:
    """Return whether Longest Road is a viable strategy for this player."""
    return _evaluate_core_special_viability(
        strategy_name="Longest Road",
        player_state=player_state,
        all_player_states=all_player_states,
        core_resource_names=("Wood", "Brick"),
        direct_all_rule="has_wood_and_brick",
        partial_with_trade_rule="has_one_road_resource_plus_favourable_trade",
        insufficient_access_rule="insufficient_wood_brick_access",
        close_to_best_rule="close_to_best_road_production",
        not_close_to_best_rule="not_close_to_best_road_production",
    )


def evaluate_largest_army_viability(
    player_state: PlayerStrategyState,
    all_player_states: Sequence[PlayerStrategyState],
) -> SpecialStrategyViability:
    """Return whether Largest Army is a viable strategy for this player."""
    return _evaluate_core_special_viability(
        strategy_name="Largest Army",
        player_state=player_state,
        all_player_states=all_player_states,
        core_resource_names=("Wheat", "Ore", "Sheep"),
        direct_all_rule="has_wheat_ore_sheep",
        partial_with_trade_rule="has_two_army_resources_plus_favourable_trade",
        insufficient_access_rule="insufficient_wheat_ore_sheep_access",
        close_to_best_rule="close_to_best_army_production",
        not_close_to_best_rule="not_close_to_best_army_production",
    )


def evaluate_special_strategy_viability(
    player_state: PlayerStrategyState,
    all_player_states: Sequence[PlayerStrategyState],
) -> Dict[str, Dict[str, Any]]:
    """Evaluate both special victory strategies for one player."""
    if not all_player_states:
        all_player_states = [player_state]

    longest_road = evaluate_longest_road_viability(player_state, all_player_states)
    largest_army = evaluate_largest_army_viability(player_state, all_player_states)

    return {
        "longest_road": longest_road.as_dict(),
        "largest_army": largest_army.as_dict(),
    }


def calculate_remaining_need(
    strategy: StrategyRequirement,
    player_state: PlayerStrategyState,
    *,
    subtract_current_roads: bool = True,
    subtract_development_cards: bool = False,
) -> StrategyRemainingNeed:
    """
    Convert one static strategy into a player-specific remaining need.

    Current settlements/cities/roads are interpreted as progress already made.
    This is resource timing, not a full legal-path plan.
    """
    current_settlements = len(player_state.settlements)
    current_cities = len(player_state.cities)
    current_total_buildings = current_settlements + current_cities

    target_total_buildings = max(strategy.target_total_buildings, strategy.cities + strategy.settlements)

    remaining_new_settlements = max(0, target_total_buildings - current_total_buildings)
    remaining_city_upgrades = max(0, strategy.cities - current_cities)

    current_roads_credit = player_state.roads_count if subtract_current_roads else 0
    remaining_roads = max(0, strategy.roads_to_build - current_roads_credit)

    dev_credit = player_state.dev_card_progress if subtract_development_cards else 0
    remaining_dev_cards = max(0, strategy.development_cards_to_buy - dev_credit)

    warnings: List[str] = []
    available_city_bases = current_settlements + remaining_new_settlements
    if remaining_city_upgrades > available_city_bases:
        warnings.append(
            "More city upgrades are required than available/current settlement bases. "
            "The resource vector is still calculated, but a legal plan should verify this."
        )

    need_vector = strategy_cost_from_components(
        new_settlements=remaining_new_settlements,
        city_upgrades=remaining_city_upgrades,
        roads=remaining_roads,
        dev_cards=remaining_dev_cards,
    )

    progress = {
        "current_settlements": current_settlements,
        "current_cities": current_cities,
        "current_total_buildings": current_total_buildings,
        "target_settlements": strategy.settlements,
        "target_cities": strategy.cities,
        "target_total_buildings": target_total_buildings,
        "current_roads_credit": current_roads_credit,
        "current_dev_card_credit": dev_credit,
        "road_progress_is_count_based": True,
        "dev_card_progress_subtracted": bool(subtract_development_cards),
    }

    return StrategyRemainingNeed(
        way_id=strategy.way_id,
        remaining_new_settlements=remaining_new_settlements,
        remaining_city_upgrades=remaining_city_upgrades,
        remaining_roads_to_build=remaining_roads,
        remaining_dev_cards_to_buy=remaining_dev_cards,
        need_vector=need_vector,
        total_cards=vector_sum(need_vector),
        progress=progress,
        warnings=tuple(warnings),
    )


# ──────────────────────────────────────────────────────────────────────────────
# EH timing and ranking
# ──────────────────────────────────────────────────────────────────────────────

def estimate_resource_requirement_time(
    *,
    current_hand: Sequence[Any],
    production_pips: Sequence[Any],
    need: Sequence[Any],
    trade_rates: Sequence[Any],
    confidence_target: float = EXPECTED_HAND_CONFIDENCE_TARGET,
    num_players: int = EXPECTED_HAND_ROLLS_PER_PLAYER_TURN,
    step: float = EXPECTED_HAND_STEP,
    max_turns: float = EXPECTED_HAND_MAX_TURNS,
    continuous_trading: bool = EXPECTED_HAND_CONTINUOUS_TRADING,
    require_confidence: bool = EXPECTED_HAND_REQUIRE_CONFIDENCE,
) -> Dict[str, Any]:
    """
    Thin wrapper around EH for arbitrary resource vectors.

    This is the main bridge needed by the 142-way strategy engine. The existing
    estimate_action_time(...) function handles named actions; this wrapper
    handles combined strategy vectors directly.
    """
    estimate = estimate_first_payable_turn(
        current_hand=clean_vector(current_hand),
        production_pips=clean_vector(production_pips),
        need=clean_vector(need),
        trade_rates=clean_int_vector(trade_rates, default=4),
        confidence_target=float(confidence_target),
        num_players=int(num_players),
        step=float(step),
        max_turns=float(max_turns),
        continuous_trading=bool(continuous_trading),
        require_confidence=bool(require_confidence),
    )

    estimate["resource_order"] = list(RESOURCE_ORDER_NAMES)
    estimate["estimator"] = "expected_hand_strategy_requirement"
    return estimate


def compute_discrete_trade_diagnostics(
    *,
    need: Sequence[Any],
    expected_hand: Sequence[Any],
    trade_rates: Sequence[Any],
    payability: Optional[Mapping[str, Any]] = None,
) -> Dict[str, Any]:
    """
    Compare optimistic continuous trading with stricter Catan bank/port trading.

    The Expected-Hand estimator may use continuous trading, where fractional
    surplus across resources can be pooled. Real Catan trading is discrete:
    each import requires enough surplus of one export resource at that resource's
    trade rate. This helper keeps both views visible.
    """
    need_v = clean_vector(need)
    hand_v = clean_vector(expected_hand)
    rates = clean_int_vector(trade_rates, default=4)

    shortages = [max(0.0, need_v[i] - hand_v[i]) for i in range(5)]
    surplus = [max(0.0, hand_v[i] - need_v[i]) for i in range(5)]

    imports_needed = sum(shortages)
    discrete_imports_available = sum(int(surplus[i] // max(1, rates[i])) for i in range(5))
    discrete_payable = imports_needed <= discrete_imports_available + 1e-9

    # User-facing margin: zero when the stricter discrete model is not payable.
    trade_margin = max(0.0, float(discrete_imports_available) - imports_needed)
    discrete_gap = max(0.0, imports_needed - float(discrete_imports_available))

    continuous_trade_margin = 0.0
    if payability:
        continuous_trade_margin = safe_float(payability.get("trades_available", 0.0), 0.0) - safe_float(
            payability.get("trades_needed", 0.0), 0.0
        )

    return {
        "short": _tuple5(shortages),
        "surplus": _tuple5(surplus),
        "imports_needed": round(imports_needed, 3),
        "discrete_imports_available": int(discrete_imports_available),
        "discrete_bank_trade_payable": bool(discrete_payable),
        "discrete_trade_gap": round(discrete_gap, 3),
        "trade_margin": round(trade_margin, 3),
        "continuous_trade_margin": round(continuous_trade_margin, 3),
        "continuous_only_trade_estimate": bool(
            imports_needed > 1e-9 and (not discrete_payable) and continuous_trade_margin >= -1e-9
        ),
        "trade_warning": (
            "continuous_only_trade_estimate"
            if imports_needed > 1e-9 and (not discrete_payable) and continuous_trade_margin >= -1e-9
            else ""
        ),
    }


def _analyze_bottlenecks(
    need: Sequence[Any],
    production_pips: Sequence[Any],
    payability: Mapping[str, Any],
    trade_rates: Optional[Sequence[Any]] = None,
) -> Tuple[Tuple[str, ...], Tuple[str, ...], str]:
    need_v = clean_vector(need)
    pips = clean_vector(production_pips)

    zero_access = [RESOURCE_ORDER_NAMES[i] for i in range(5) if need_v[i] > 1e-9 and pips[i] <= 1e-9]

    imports = clean_vector(payability.get("imports_received", [0, 0, 0, 0, 0]))
    exports = clean_vector(payability.get("exports_used", [0, 0, 0, 0, 0]))
    trade_imports = [RESOURCE_ORDER_NAMES[i] for i in range(5) if imports[i] > 1e-9]
    trade_exports = [RESOURCE_ORDER_NAMES[i] for i in range(5) if exports[i] > 1e-9]

    dependency: List[str] = []
    if trade_imports:
        dependency.append("imports: " + ", ".join(trade_imports))
    if trade_exports:
        dependency.append("exports: " + ", ".join(trade_exports))

    rates = clean_int_vector(trade_rates, default=4) if trade_rates is not None else [4, 4, 4, 4, 4]
    trade_label = "bank trades" if min(rates) >= 4 else "bank/port trades"

    if zero_access:
        reason = "trade-dependent for " + ", ".join(zero_access)
    elif trade_imports:
        reason = f"uses {trade_label} for " + ", ".join(trade_imports)
    else:
        reason = "mostly direct production/hand fit"

    return tuple(zero_access), tuple(dependency), reason


def _row_sort_key(row: StrategyTimingRow) -> Tuple[Any, ...]:
    trade_diag = dict(row.trade_diagnostics or {})
    continuous_only = bool(trade_diag.get("continuous_only_trade_estimate", False))
    discrete_gap = safe_float(trade_diag.get("discrete_trade_gap", 0.0), 0.0)
    return (
        not row.found,
        continuous_only,
        discrete_gap,
        finite_or_9999(row.turns),
        -safe_float(row.confidence, 0.0),
        row.remaining.total_cards,
        row.way_id,
    )


def rank_strategies_for_player_state(
    player_state: PlayerStrategyState,
    requirements: Optional[Sequence[StrategyRequirement]] = None,
    *,
    requirements_path: Optional[os.PathLike[str] | str] = None,
    top_n: Optional[int] = 3,
    include_all: bool = True,
    include_debug: bool = False,
    confidence_target: float = EXPECTED_HAND_CONFIDENCE_TARGET,
    num_players: Optional[int] = None,
    step: float = EXPECTED_HAND_STEP,
    max_turns: float = EXPECTED_HAND_MAX_TURNS,
    continuous_trading: bool = EXPECTED_HAND_CONTINUOUS_TRADING,
    require_confidence: bool = EXPECTED_HAND_REQUIRE_CONFIDENCE,
    subtract_current_roads: bool = True,
    subtract_development_cards: bool = False,
    all_player_states: Optional[Sequence[PlayerStrategyState]] = None,
    exclude_non_viable_special_strategies: bool = DEFAULT_EXCLUDE_NON_VIABLE_SPECIAL_STRATEGIES,
) -> Dict[str, Any]:
    """Rank all 142 strategies from an already-built player state.

    This entry point is used by action_planner.py Stage 3A. It lets a
    hypothetical after-action player state be ranked without mutating game,
    board, or player objects.
    """
    if requirements is None:
        requirements = load_strategy_requirements(requirements_path)

    strategy_list = list(requirements)

    if num_players is None:
        num_players = EXPECTED_HAND_ROLLS_PER_PLAYER_TURN

    if all_player_states is None:
        all_player_states = [player_state]
    else:
        all_player_states = list(all_player_states)
        if not all_player_states:
            all_player_states = [player_state]

    special_viability = evaluate_special_strategy_viability(player_state, all_player_states)

    # Multiple ways can collapse to the same remaining resource vector. EH only
    # needs to run once per unique need vector for this player state.
    estimate_by_need: Dict[Tuple[float, float, float, float, float], Dict[str, Any]] = {}
    rows: List[StrategyTimingRow] = []
    rejected_special: List[Dict[str, Any]] = []

    for strategy in strategy_list:
        strategy_special_viability: Dict[str, Any] = {}
        failed_special_reasons: List[str] = []
        failed_special_rules: List[str] = []

        if strategy.longest_road:
            lr = special_viability.get("longest_road", {}) or {}
            strategy_special_viability["longest_road"] = lr
            if exclude_non_viable_special_strategies and not bool(lr.get("viable", False)):
                failed_special_reasons.append(str(lr.get("reason", "Longest Road not viable")))
                failed_special_rules.extend(str(rule) for rule in lr.get("failed_rules", []) or [])

        if strategy.biggest_army:
            la = special_viability.get("largest_army", {}) or {}
            strategy_special_viability["largest_army"] = la
            if exclude_non_viable_special_strategies and not bool(la.get("viable", False)):
                failed_special_reasons.append(str(la.get("reason", "Largest Army not viable")))
                failed_special_rules.extend(str(rule) for rule in la.get("failed_rules", []) or [])

        if failed_special_reasons:
            rejected_special.append(
                {
                    "way_id": strategy.way_id,
                    "tags": list(strategy.tags),
                    "reason": " | ".join(failed_special_reasons),
                    "failed_rules": sorted(set(failed_special_rules)),
                    "special_viability": strategy_special_viability,
                }
            )
            continue

        remaining = calculate_remaining_need(
            strategy,
            player_state,
            subtract_current_roads=subtract_current_roads,
            subtract_development_cards=subtract_development_cards,
        )

        need_key = _tuple5(remaining.need_vector)
        if need_key not in estimate_by_need:
            estimate_by_need[need_key] = estimate_resource_requirement_time(
                current_hand=player_state.current_hand,
                production_pips=player_state.production_pips,
                need=need_key,
                trade_rates=player_state.trade_rates,
                confidence_target=confidence_target,
                num_players=int(num_players),
                step=step,
                max_turns=max_turns,
                continuous_trading=continuous_trading,
                require_confidence=require_confidence,
            )

        estimate = estimate_by_need[need_key]
        turns = finite_or_9999(estimate.get("turns"))
        found = bool(estimate.get("found", False))
        confidence = safe_float(estimate.get("confidence"), 0.0)
        payability = estimate.get("payability", {}) or {}
        expected_hand = _tuple5(estimate.get("expected_hand", [0, 0, 0, 0, 0]))
        trade_diagnostics = compute_discrete_trade_diagnostics(
            need=need_key,
            expected_hand=expected_hand,
            trade_rates=player_state.trade_rates,
            payability=payability,
        )
        bottlenecks, dependency, reason = _analyze_bottlenecks(
            need_key,
            player_state.production_pips,
            payability,
            player_state.trade_rates,
        )

        rows.append(
            StrategyTimingRow(
                way_id=strategy.way_id,
                rank=0,
                turns=turns,
                found=found,
                confidence=confidence,
                confidence_label=str(estimate.get("confidence_label", "")),
                need_vector=need_key,
                expected_hand=expected_hand,
                trade_rates=player_state.trade_rates,
                tags=tuple(strategy.tags),
                reason=reason,
                bottlenecks=bottlenecks,
                trade_dependency=dependency,
                strategy=strategy,
                remaining=remaining,
                estimate=estimate if include_debug else {},
                special_viability=strategy_special_viability,
                trade_diagnostics=trade_diagnostics,
            )
        )

    rows.sort(key=_row_sort_key)
    ranked_rows: List[StrategyTimingRow] = []
    for idx, row in enumerate(rows, start=1):
        ranked_rows.append(
            StrategyTimingRow(
                way_id=row.way_id,
                rank=idx,
                turns=row.turns,
                found=row.found,
                confidence=row.confidence,
                confidence_label=row.confidence_label,
                need_vector=row.need_vector,
                expected_hand=row.expected_hand,
                trade_rates=row.trade_rates,
                tags=row.tags,
                reason=row.reason,
                bottlenecks=row.bottlenecks,
                trade_dependency=row.trade_dependency,
                strategy=row.strategy,
                remaining=row.remaining,
                estimate=row.estimate,
                special_viability=row.special_viability,
                trade_diagnostics=row.trade_diagnostics,
            )
        )

    n = int(top_n) if top_n is not None else 3
    top_rows = ranked_rows[: max(0, n)]

    best_non_trade = next((r for r in ranked_rows if r.found and not r.trade_dependency), None)
    best_high_confidence = next((r for r in ranked_rows if r.found and r.confidence_label in {"exact", "high"}), None)

    output: Dict[str, Any] = {
        "player": player_state.as_dict(),
        "special_strategy_viability": special_viability,
        "settings": {
            "top_n": top_n,
            "include_all": include_all,
            "include_debug": include_debug,
            "confidence_target": confidence_target,
            "num_players": int(num_players),
            "step": step,
            "max_turns": max_turns,
            "continuous_trading": continuous_trading,
            "require_confidence": require_confidence,
            "subtract_current_roads": subtract_current_roads,
            "subtract_development_cards": subtract_development_cards,
            "exclude_non_viable_special_strategies": bool(exclude_non_viable_special_strategies),
            "special_close_to_best_ratio": SPECIAL_CLOSE_TO_BEST_RATIO,
            "special_access_pip_threshold": SPECIAL_ACCESS_PIP_THRESHOLD,
            "special_excess_pips_threshold": SPECIAL_EXCESS_PIPS_THRESHOLD,
            "special_favourable_trade_rate": SPECIAL_FAVOURABLE_TRADE_RATE,
            "unique_need_vectors_evaluated": len(estimate_by_need),
            "strategy_count": len(strategy_list),
            "excluded_special_strategy_count": len(rejected_special),
            "rte_available": _RTE_AVAILABLE,
        },
        "top_strategies": [row.as_dict(include_debug=include_debug) for row in top_rows],
        "best_non_trade_dependent_strategy": (
            best_non_trade.as_dict(include_debug=include_debug) if best_non_trade else None
        ),
        "best_high_confidence_strategy": (
            best_high_confidence.as_dict(include_debug=include_debug) if best_high_confidence else None
        ),
    }

    if include_all:
        output["all_strategies"] = [row.as_dict(include_debug=include_debug) for row in ranked_rows]

    if include_debug:
        output["rejected_special_strategies"] = rejected_special

    warnings = []
    for strategy in strategy_list:
        for warning in strategy.validation_warnings:
            warnings.append({"way_id": strategy.way_id, "warning": warning})
    if warnings:
        output["loader_warnings"] = warnings

    return output


def rank_strategies_for_player(
    board: Any,
    player: Any,
    requirements: Optional[Sequence[StrategyRequirement]] = None,
    *,
    requirements_path: Optional[os.PathLike[str] | str] = None,
    top_n: Optional[int] = 3,
    include_all: bool = True,
    include_debug: bool = False,
    confidence_target: float = EXPECTED_HAND_CONFIDENCE_TARGET,
    num_players: Optional[int] = None,
    step: float = EXPECTED_HAND_STEP,
    max_turns: float = EXPECTED_HAND_MAX_TURNS,
    continuous_trading: bool = EXPECTED_HAND_CONTINUOUS_TRADING,
    require_confidence: bool = EXPECTED_HAND_REQUIRE_CONFIDENCE,
    subtract_current_roads: bool = True,
    subtract_development_cards: bool = False,
    all_player_states: Optional[Sequence[PlayerStrategyState]] = None,
    exclude_non_viable_special_strategies: bool = DEFAULT_EXCLUDE_NON_VIABLE_SPECIAL_STRATEGIES,
) -> Dict[str, Any]:
    """Rank all 142 strategies for one player using EH timing."""
    if requirements is None:
        requirements = load_strategy_requirements(requirements_path)

    strategy_list = list(requirements)
    player_state = build_player_strategy_state(board, player)

    game = getattr(player, "game", None)
    players = getattr(game, "players", None)

    if num_players is None:
        try:
            num_players = max(1, len(players))
        except Exception:
            num_players = EXPECTED_HAND_ROLLS_PER_PLAYER_TURN

    if all_player_states is None:
        try:
            all_player_states = [build_player_strategy_state(board, p) for p in players]
        except Exception:
            all_player_states = [player_state]
    else:
        all_player_states = list(all_player_states)
        if not all_player_states:
            all_player_states = [player_state]

    special_viability = evaluate_special_strategy_viability(player_state, all_player_states)

    # Multiple ways can collapse to the same remaining resource vector. EH only
    # needs to run once per unique need vector for this player state.
    estimate_by_need: Dict[Tuple[float, float, float, float, float], Dict[str, Any]] = {}
    rows: List[StrategyTimingRow] = []
    rejected_special: List[Dict[str, Any]] = []

    for strategy in strategy_list:
        strategy_special_viability: Dict[str, Any] = {}
        failed_special_reasons: List[str] = []
        failed_special_rules: List[str] = []

        if strategy.longest_road:
            lr = special_viability.get("longest_road", {}) or {}
            strategy_special_viability["longest_road"] = lr
            if exclude_non_viable_special_strategies and not bool(lr.get("viable", False)):
                failed_special_reasons.append(str(lr.get("reason", "Longest Road not viable")))
                failed_special_rules.extend(str(rule) for rule in lr.get("failed_rules", []) or [])

        if strategy.biggest_army:
            la = special_viability.get("largest_army", {}) or {}
            strategy_special_viability["largest_army"] = la
            if exclude_non_viable_special_strategies and not bool(la.get("viable", False)):
                failed_special_reasons.append(str(la.get("reason", "Largest Army not viable")))
                failed_special_rules.extend(str(rule) for rule in la.get("failed_rules", []) or [])

        if failed_special_reasons:
            rejected_special.append(
                {
                    "way_id": strategy.way_id,
                    "tags": list(strategy.tags),
                    "reason": " | ".join(failed_special_reasons),
                    "failed_rules": sorted(set(failed_special_rules)),
                    "special_viability": strategy_special_viability,
                }
            )
            continue

        remaining = calculate_remaining_need(
            strategy,
            player_state,
            subtract_current_roads=subtract_current_roads,
            subtract_development_cards=subtract_development_cards,
        )

        need_key = _tuple5(remaining.need_vector)
        if need_key not in estimate_by_need:
            estimate_by_need[need_key] = estimate_resource_requirement_time(
                current_hand=player_state.current_hand,
                production_pips=player_state.production_pips,
                need=need_key,
                trade_rates=player_state.trade_rates,
                confidence_target=confidence_target,
                num_players=int(num_players),
                step=step,
                max_turns=max_turns,
                continuous_trading=continuous_trading,
                require_confidence=require_confidence,
            )

        estimate = estimate_by_need[need_key]
        turns = finite_or_9999(estimate.get("turns"))
        found = bool(estimate.get("found", False))
        confidence = safe_float(estimate.get("confidence"), 0.0)
        payability = estimate.get("payability", {}) or {}
        expected_hand = _tuple5(estimate.get("expected_hand", [0, 0, 0, 0, 0]))
        trade_diagnostics = compute_discrete_trade_diagnostics(
            need=need_key,
            expected_hand=expected_hand,
            trade_rates=player_state.trade_rates,
            payability=payability,
        )
        bottlenecks, dependency, reason = _analyze_bottlenecks(
            need_key,
            player_state.production_pips,
            payability,
            player_state.trade_rates,
        )

        rows.append(
            StrategyTimingRow(
                way_id=strategy.way_id,
                rank=0,
                turns=turns,
                found=found,
                confidence=confidence,
                confidence_label=str(estimate.get("confidence_label", "")),
                need_vector=need_key,
                expected_hand=expected_hand,
                trade_rates=player_state.trade_rates,
                tags=tuple(strategy.tags),
                reason=reason,
                bottlenecks=bottlenecks,
                trade_dependency=dependency,
                strategy=strategy,
                remaining=remaining,
                estimate=estimate if include_debug else {},
                special_viability=strategy_special_viability,
                trade_diagnostics=trade_diagnostics,
            )
        )

    rows.sort(key=_row_sort_key)
    ranked_rows: List[StrategyTimingRow] = []
    for idx, row in enumerate(rows, start=1):
        ranked_rows.append(
            StrategyTimingRow(
                way_id=row.way_id,
                rank=idx,
                turns=row.turns,
                found=row.found,
                confidence=row.confidence,
                confidence_label=row.confidence_label,
                need_vector=row.need_vector,
                expected_hand=row.expected_hand,
                trade_rates=row.trade_rates,
                tags=row.tags,
                reason=row.reason,
                bottlenecks=row.bottlenecks,
                trade_dependency=row.trade_dependency,
                strategy=row.strategy,
                remaining=row.remaining,
                estimate=row.estimate,
                special_viability=row.special_viability,
                trade_diagnostics=row.trade_diagnostics,
            )
        )

    n = int(top_n) if top_n is not None else 3
    top_rows = ranked_rows[: max(0, n)]

    best_non_trade = next((r for r in ranked_rows if r.found and not r.trade_dependency), None)
    best_high_confidence = next((r for r in ranked_rows if r.found and r.confidence_label in {"exact", "high"}), None)

    output: Dict[str, Any] = {
        "player": player_state.as_dict(),
        "special_strategy_viability": special_viability,
        "settings": {
            "top_n": top_n,
            "include_all": include_all,
            "include_debug": include_debug,
            "confidence_target": confidence_target,
            "num_players": int(num_players),
            "step": step,
            "max_turns": max_turns,
            "continuous_trading": continuous_trading,
            "require_confidence": require_confidence,
            "subtract_current_roads": subtract_current_roads,
            "subtract_development_cards": subtract_development_cards,
            "exclude_non_viable_special_strategies": bool(exclude_non_viable_special_strategies),
            "special_close_to_best_ratio": SPECIAL_CLOSE_TO_BEST_RATIO,
            "special_access_pip_threshold": SPECIAL_ACCESS_PIP_THRESHOLD,
            "special_excess_pips_threshold": SPECIAL_EXCESS_PIPS_THRESHOLD,
            "special_favourable_trade_rate": SPECIAL_FAVOURABLE_TRADE_RATE,
            "unique_need_vectors_evaluated": len(estimate_by_need),
            "strategy_count": len(strategy_list),
            "excluded_special_strategy_count": len(rejected_special),
            "rte_available": _RTE_AVAILABLE,
        },
        "top_strategies": [row.as_dict(include_debug=include_debug) for row in top_rows],
        "best_non_trade_dependent_strategy": (
            best_non_trade.as_dict(include_debug=include_debug) if best_non_trade else None
        ),
        "best_high_confidence_strategy": (
            best_high_confidence.as_dict(include_debug=include_debug) if best_high_confidence else None
        ),
    }

    if include_all:
        output["all_strategies"] = [row.as_dict(include_debug=include_debug) for row in ranked_rows]

    if include_debug:
        output["rejected_special_strategies"] = rejected_special

    warnings = []
    for strategy in strategy_list:
        for warning in strategy.validation_warnings:
            warnings.append({"way_id": strategy.way_id, "warning": warning})
    if warnings:
        output["loader_warnings"] = warnings

    return output

# ──────────────────────────────────────────────────────────────────────────────
# Report-only player trade opportunities
# ──────────────────────────────────────────────────────────────────────────────

def _report_player_id_key(by_player: Mapping[Any, Any], player_id: int) -> Optional[Any]:
    """Return the key used in by_player for this player id, supporting int/string keys."""
    if player_id in by_player:
        return player_id
    text_key = str(player_id)
    if text_key in by_player:
        return text_key
    return None


def _player_victory_points(player: Any) -> int:
    """
    Defensive victory-point lookup.

    Your Player object has both points and victory_points in different contexts.
    For trading, use the larger known value.
    """
    return max(
        safe_int(getattr(player, "victory_points", 0), 0),
        safe_int(getattr(player, "points", 0), 0),
    )


def _expected_hand_from_report_player(
    player_report: Mapping[str, Any],
    *,
    own_turns: float,
) -> Tuple[float, float, float, float, float]:
    """
    Lightweight expected hand for trade diagnostics.

    Uses:
        current_hand + production_pips * expected_dice_rolls / 36

    This is report-only and does not mutate anything.
    """
    player_info = player_report.get("player", {}) or {}
    settings = player_report.get("settings", {}) or {}

    current_hand = clean_vector(player_info.get("current_hand", [0, 0, 0, 0, 0]))
    production_pips = clean_vector(player_info.get("production_pips", [0, 0, 0, 0, 0]))

    num_players = max(1, safe_int(settings.get("num_players", NUM_PLAYERS), NUM_PLAYERS))
    expected_dice_rolls = max(0.0, safe_float(own_turns, 0.0)) * num_players

    expected = [
        current_hand[i] + production_pips[i] * expected_dice_rolls / 36.0
        for i in range(5)
    ]
    return _tuple5(expected)


def _positive_vector_diff(a: Sequence[Any], b: Sequence[Any]) -> Tuple[float, float, float, float, float]:
    """Return max(a - b, 0) per resource."""
    av = clean_vector(a)
    bv = clean_vector(b)
    return _tuple5([max(0.0, av[i] - bv[i]) for i in range(5)])


def _shortage_vector(need: Sequence[Any], hand: Sequence[Any]) -> Tuple[float, float, float, float, float]:
    """Return max(need - hand, 0) per resource."""
    return _positive_vector_diff(need, hand)


def _resource_names_from_vector(values: Sequence[Any], *, threshold: float = 1e-9) -> List[str]:
    vec = clean_vector(values)
    return [RESOURCE_ORDER_NAMES[i] for i in range(5) if vec[i] > threshold]


def _table_resource_scarcity_from_report(
    report: Mapping[str, Any],
    *,
    weak_access_pips_threshold: float = 3.0,
    table_scarce_total_pips_threshold: float = 6.0,
) -> Dict[str, Dict[str, Any]]:
    """
    Resource scarcity diagnostics based on production pips.

    Important distinction:
        table_scarce:
            The resource is scarce at table level.

        scarce_for_players:
            The resource is scarce for specific players, even if it is not
            scarce at table level.

    Example:
        Brick total_pips = 11
        Player 1 = 0, Player 2 = 0, Player 3 = 4, Player 4 = 7

        table_scarce = False
        scarce_for_players = [1, 2]
        players_without_access = [1, 2]
        monopoly_risk = True
    """
    by_player = report.get("by_player", {}) or {}

    scarcity: Dict[str, Dict[str, Any]] = {}

    for idx, resource_name in enumerate(RESOURCE_ORDER_NAMES):
        player_pips: Dict[int, float] = {}
        total_pips = 0.0

        players_with_access: List[int] = []
        players_without_access: List[int] = []
        players_with_weak_access: List[int] = []
        players_with_good_access: List[int] = []
        scarce_for_players: List[int] = []

        for _, player_report in by_player.items():
            player_info = player_report.get("player", {}) or {}
            pid = safe_int(player_info.get("player_id", 0), 0)
            if pid <= 0:
                continue

            pips = clean_vector(player_info.get("production_pips", [0, 0, 0, 0, 0]))
            value = safe_float(pips[idx], 0.0)

            player_pips[pid] = round(value, 3)
            total_pips += value

            if value <= 1e-9:
                players_without_access.append(pid)
                scarce_for_players.append(pid)
            else:
                players_with_access.append(pid)

                if value < weak_access_pips_threshold:
                    players_with_weak_access.append(pid)
                    scarce_for_players.append(pid)
                else:
                    players_with_good_access.append(pid)

        players_with_access_count = len(players_with_access)

        table_scarce = (
            players_with_access_count <= 1
            or total_pips <= table_scarce_total_pips_threshold
        )

        # Monopoly risk means only one player has meaningful/good access, or
        # at least half the table has no access at all.
        monopoly_risk = (
            len(players_with_good_access) <= 1
            or len(players_without_access) >= max(2, len(player_pips) // 2)
        )

        scarcity[resource_name] = {
            "total_pips": round(total_pips, 3),
            "players_with_access": players_with_access_count,
            "player_pips": player_pips,

            # New clearer fields.
            "table_scarce": bool(table_scarce),
            "scarce_for_players": sorted(set(scarce_for_players)),
            "players_without_access": sorted(set(players_without_access)),
            "players_with_weak_access": sorted(set(players_with_weak_access)),
            "players_with_good_access": sorted(set(players_with_good_access)),
            "monopoly_risk": bool(monopoly_risk),

            # Backward-compatible alias.
            "scarce": bool(table_scarce),
        }

    return scarcity


def _row_need_vector(row: Mapping[str, Any]) -> Tuple[float, float, float, float, float]:
    """Read a strategy row need vector defensively."""
    return _tuple5(row.get("need_vector", [0, 0, 0, 0, 0]))


def _row_expected_own_turns(row: Mapping[str, Any]) -> float:
    """Read a strategy row timing defensively."""
    return finite_or_9999(row.get("expected_own_turns", row.get("turns", INFINITE_TURNS)))


def _best_reference_need_for_player(player_report: Mapping[str, Any]) -> Tuple[float, float, float, float, float]:
    """
    Use the player's current #1 top strategy as their reference need.

    This answers:
        What does this player probably still need for their own best route?
    """
    top = list(player_report.get("top_strategies", []) or [])
    if not top:
        return _tuple5([0, 0, 0, 0, 0])
    return _row_need_vector(top[0])


def _find_trade_opportunities_for_strategy(
    *,
    from_player: Any,
    from_report: Mapping[str, Any],
    strategy_row: Mapping[str, Any],
    counterparty: Any,
    counterparty_report: Mapping[str, Any],
    table_scarcity: Mapping[str, Mapping[str, Any]],
    min_trade_surplus_cards: float,
    max_trade_victory_points: int,
    max_opportunities: int,
) -> List[Dict[str, Any]]:
    """
    Find report-only player-trade opportunities for one top strategy row.

    Player A gives a resource useful/scarce for Player B.
    Player A receives a resource that reduces A's shortage and is expected
    surplus for Player B.

    This does not change timing, hand, board, or need vectors.
    """
    from_player_id = safe_int(getattr(from_player, "id", 0), 0)
    counterparty_id = safe_int(getattr(counterparty, "id", 0), 0)

    if from_player_id == counterparty_id:
        return []

    from_vp = _player_victory_points(from_player)
    counterparty_vp = _player_victory_points(counterparty)

    if from_vp > max_trade_victory_points or counterparty_vp > max_trade_victory_points:
        return []

    own_turns = _row_expected_own_turns(strategy_row)
    if own_turns >= INFINITE_TURNS:
        return []

    # A's strategy need and expected shortage.
    a_need = _row_need_vector(strategy_row)
    a_expected_hand = _expected_hand_from_report_player(from_report, own_turns=own_turns)
    a_shortage = _shortage_vector(a_need, a_expected_hand)
    a_surplus = _positive_vector_diff(a_expected_hand, a_need)

    # B's reference need from B's own best strategy.
    b_reference_need = _best_reference_need_for_player(counterparty_report)
    b_expected_hand = _expected_hand_from_report_player(counterparty_report, own_turns=own_turns)
    b_shortage = _shortage_vector(b_reference_need, b_expected_hand)
    b_surplus = _positive_vector_diff(b_expected_hand, b_reference_need)

    b_info = counterparty_report.get("player", {}) or {}
    b_pips = clean_vector(b_info.get("production_pips", [0, 0, 0, 0, 0]))

    opportunities: List[Dict[str, Any]] = []

    for receive_idx, receive_name in enumerate(RESOURCE_ORDER_NAMES):
        # A only wants resources that reduce A's remaining shortage.
        if a_shortage[receive_idx] <= 1e-9:
            continue

        # B must be expected to have the received resource in surplus.
        if b_surplus[receive_idx] < min_trade_surplus_cards:
            continue

        for give_idx, give_name in enumerate(RESOURCE_ORDER_NAMES):
            if give_idx == receive_idx:
                continue

            # A must be expected to have the offered resource in surplus.
            if a_surplus[give_idx] < 1.0:
                continue

            # B must value the offered resource:
            #   - it reduces B's shortage, or
            #   - B has weak/no production and the table says this resource is scarce.
            give_is_shortage_for_b = b_shortage[give_idx] > 1e-9
            scarcity_info = table_scarcity.get(give_name, {}) or {}
            scarce_for_players = set(safe_int(pid, 0) for pid in scarcity_info.get("scarce_for_players", []))

            give_is_scarce_for_b = counterparty_id in scarce_for_players

            if not (give_is_shortage_for_b or give_is_scarce_for_b):
                continue

            confidence = min(
                1.0,
                max(
                    0.1,
                    b_surplus[receive_idx] / max(min_trade_surplus_cards, 1e-9),
                    a_surplus[give_idx] / max(min_trade_surplus_cards, 1e-9),
                ),
            )

            status = "strong" if confidence >= 0.85 else "possible"

            reason_parts = [
                f"Player {from_player_id} needs {receive_name}.",
                f"Player {counterparty_id} is expected to have surplus {receive_name}.",
                f"Player {from_player_id} can offer surplus {give_name}.",
            ]

            if give_is_shortage_for_b:
                reason_parts.append(f"{give_name} reduces Player {counterparty_id}'s shortage.")
            elif give_is_scarce_for_b:
                reason_parts.append(f"{give_name} is scarce for Player {counterparty_id}.")
            reason_parts.append(
                f"Trade allowed because both players have <= {max_trade_victory_points} victory points."
            )

            opportunities.append(
                {
                    "from_player_id": from_player_id,
                    "counterparty_player_id": counterparty_id,
                    "give": give_name,
                    "receive": receive_name,
                    "status": status,
                    "confidence": round(confidence, 3),
                    "reason": " ".join(reason_parts),
                    "from_player_victory_points": from_vp,
                    "counterparty_victory_points": counterparty_vp,
                    "from_player_expected_surplus_give": round(a_surplus[give_idx], 3),
                    "counterparty_expected_surplus_receive": round(b_surplus[receive_idx], 3),
                    "counterparty_shortage_for_give": round(b_shortage[give_idx], 3),
                    "strategy_way_id": strategy_row.get("way_id"),
                    "strategy_rank": strategy_row.get("rank"),
                }
            )

    opportunities.sort(
        key=lambda item: (
            0 if item.get("status") == "strong" else 1,
            -safe_float(item.get("confidence", 0.0), 0.0),
            str(item.get("receive", "")),
            str(item.get("give", "")),
        )
    )

    return opportunities[: max(0, int(max_opportunities))]


def add_player_trade_opportunities_to_report(
    game: Any,
    report: Dict[str, Any],
    *,
    min_trade_surplus_cards: float = DEFAULT_MIN_TRADE_SURPLUS_CARDS,
    max_trade_victory_points: int = DEFAULT_MAX_PLAYER_TRADE_VICTORY_POINTS,
    max_opportunities_per_strategy: int = DEFAULT_MAX_TRADE_OPPORTUNITIES_PER_STRATEGY,
) -> Dict[str, Any]:
    """
    Add report-only player_trade_opportunities to each top strategy row.

    This function intentionally:
        - does not mutate game/player/board state
        - does not mutate need vectors
        - does not change EH timing or ranking
    """
    by_player = report.get("by_player", {}) or {}
    table_scarcity = _table_resource_scarcity_from_report(report)

    player_by_id = {
        safe_int(getattr(player, "id", 0), 0): player
        for player in getattr(game, "players", []) or []
    }

    report["player_trade_settings"] = {
        "enabled": True,
        "report_only": True,
        "max_trade_victory_points": max_trade_victory_points,
        "min_trade_surplus_cards": min_trade_surplus_cards,
        "max_opportunities_per_strategy": max_opportunities_per_strategy,
        "table_resource_scarcity": table_scarcity,
    }

    for player_id, player in player_by_id.items():
        player_key = _report_player_id_key(by_player, player_id)
        if player_key is None:
            continue

        player_report = by_player[player_key]
        top_rows = list(player_report.get("top_strategies", []) or [])

        for row in top_rows:
            row["player_trade_opportunities"] = []

            for counterparty_id, counterparty in player_by_id.items():
                if counterparty_id == player_id:
                    continue

                counterparty_key = _report_player_id_key(by_player, counterparty_id)
                if counterparty_key is None:
                    continue

                counterparty_report = by_player[counterparty_key]

                row["player_trade_opportunities"].extend(
                    _find_trade_opportunities_for_strategy(
                        from_player=player,
                        from_report=player_report,
                        strategy_row=row,
                        counterparty=counterparty,
                        counterparty_report=counterparty_report,
                        table_scarcity=table_scarcity,
                        min_trade_surplus_cards=min_trade_surplus_cards,
                        max_trade_victory_points=max_trade_victory_points,
                        max_opportunities=max_opportunities_per_strategy,
                    )
                )

            row["player_trade_opportunities"].sort(
                key=lambda item: (
                    0 if item.get("status") == "strong" else 1,
                    -safe_float(item.get("confidence", 0.0), 0.0),
                    safe_int(item.get("counterparty_player_id", 999), 999),
                )
            )

            row["player_trade_opportunities"] = row["player_trade_opportunities"][
                : max(0, int(max_opportunities_per_strategy))
            ]

    return report


def build_strategy_timing_report(
    game: Any,
    requirements: Optional[Sequence[StrategyRequirement]] = None,
    *,
    requirements_path: Optional[os.PathLike[str] | str] = None,
    top_n: int = 3,
    include_all: bool = False,
    include_debug: bool = False,
    player_ids: Optional[Iterable[int]] = None,

    # Report-only player-trade diagnostics.
    include_player_trade_opportunities: bool = DEFAULT_INCLUDE_PLAYER_TRADE_OPPORTUNITIES,
    max_trade_victory_points: int = DEFAULT_MAX_PLAYER_TRADE_VICTORY_POINTS,
    min_trade_surplus_cards: float = DEFAULT_MIN_TRADE_SURPLUS_CARDS,
    max_trade_opportunities_per_strategy: int = DEFAULT_MAX_TRADE_OPPORTUNITIES_PER_STRATEGY,

    **timing_kwargs: Any,
) -> Dict[str, Any]:
    """Build a JSON-friendly strategy timing report for one or more players."""

    # Extra safety: if these arrive through **timing_kwargs anyway,
    # consume them here so rank_strategies_for_player(...) never sees them.
    include_player_trade_opportunities = bool(
        timing_kwargs.pop(
            "include_player_trade_opportunities",
            include_player_trade_opportunities,
        )
    )
    max_trade_victory_points = safe_int(
        timing_kwargs.pop(
            "max_trade_victory_points",
            max_trade_victory_points,
        ),
        max_trade_victory_points,
    )
    min_trade_surplus_cards = safe_float(
        timing_kwargs.pop(
            "min_trade_surplus_cards",
            min_trade_surplus_cards,
        ),
        min_trade_surplus_cards,
    )
    max_trade_opportunities_per_strategy = safe_int(
        timing_kwargs.pop(
            "max_trade_opportunities_per_strategy",
            max_trade_opportunities_per_strategy,
        ),
        max_trade_opportunities_per_strategy,
    )

    if requirements is None:
        requirements = load_strategy_requirements(requirements_path)

    selected_ids = set(int(x) for x in player_ids) if player_ids is not None else None

    players = []
    for player in getattr(game, "players", []) or []:
        pid = safe_int(getattr(player, "id", 0), 0)
        if selected_ids is not None and pid not in selected_ids:
            continue
        players.append(player)

    board = getattr(game, "board")
    all_player_states = [build_player_strategy_state(board, player) for player in players]

    by_player = {}
    for player in players:
        pid = safe_int(getattr(player, "id", 0), 0)
        by_player[pid] = rank_strategies_for_player(
            board,
            player,
            requirements=requirements,
            top_n=top_n,
            include_all=include_all,
            include_debug=include_debug,
            all_player_states=all_player_states,
            **timing_kwargs,
        )

    report = {
        "round": getattr(game, "round", None),
        "turn": getattr(game, "turn", None),
        "phase": getattr(game, "phase", None),
        "strategy_count": len(list(requirements)),
        "top_n": top_n,
        "resource_order": list(RESOURCE_ORDER_NAMES),
        "special_strategy_settings": {
            "exclude_non_viable_special_strategies_default": DEFAULT_EXCLUDE_NON_VIABLE_SPECIAL_STRATEGIES,
            "close_to_best_ratio": SPECIAL_CLOSE_TO_BEST_RATIO,
            "access_pip_threshold": SPECIAL_ACCESS_PIP_THRESHOLD,
            "excess_pips_threshold": SPECIAL_EXCESS_PIPS_THRESHOLD,
            "favourable_trade_rate": SPECIAL_FAVOURABLE_TRADE_RATE,
        },
        "by_player": by_player,
    }

    if include_player_trade_opportunities:
        add_player_trade_opportunities_to_report(
            game,
            report,
            min_trade_surplus_cards=min_trade_surplus_cards,
            max_trade_victory_points=max_trade_victory_points,
            max_opportunities_per_strategy=max_trade_opportunities_per_strategy,
        )

    return report


class StrategyTimingEngine:
    """Reusable engine that loads requirements once and ranks players cheaply."""

    def __init__(
        self,
        requirements_path: Optional[os.PathLike[str] | str] = None,
        *,
        requirements: Optional[Sequence[StrategyRequirement]] = None,
        base_dir: Optional[os.PathLike[str] | str] = None,
        sheet_name: Optional[str] = None,
        confidence_target: float = EXPECTED_HAND_CONFIDENCE_TARGET,
        num_players: Optional[int] = None,
        step: float = EXPECTED_HAND_STEP,
        max_turns: float = EXPECTED_HAND_MAX_TURNS,
        continuous_trading: bool = EXPECTED_HAND_CONTINUOUS_TRADING,
        require_confidence: bool = EXPECTED_HAND_REQUIRE_CONFIDENCE,
        subtract_current_roads: bool = True,
        subtract_development_cards: bool = False,
        exclude_non_viable_special_strategies: bool = DEFAULT_EXCLUDE_NON_VIABLE_SPECIAL_STRATEGIES,
    ) -> None:
        self.requirements_path = requirements_path
        self.requirements = list(
            requirements
            if requirements is not None
            else load_strategy_requirements(requirements_path, base_dir=base_dir, sheet_name=sheet_name)
        )
        self.confidence_target = confidence_target
        self.num_players = num_players
        self.step = step
        self.max_turns = max_turns
        self.continuous_trading = continuous_trading
        self.require_confidence = require_confidence
        self.subtract_current_roads = subtract_current_roads
        self.subtract_development_cards = subtract_development_cards
        self.exclude_non_viable_special_strategies = exclude_non_viable_special_strategies

    def _timing_kwargs(self, overrides: Mapping[str, Any]) -> Dict[str, Any]:
        kwargs = {
            "confidence_target": self.confidence_target,
            "num_players": self.num_players,
            "step": self.step,
            "max_turns": self.max_turns,
            "continuous_trading": self.continuous_trading,
            "require_confidence": self.require_confidence,
            "subtract_current_roads": self.subtract_current_roads,
            "subtract_development_cards": self.subtract_development_cards,
            "exclude_non_viable_special_strategies": self.exclude_non_viable_special_strategies,
        }
        kwargs.update({k: v for k, v in overrides.items() if v is not None})
        return kwargs

    def rank_for_player(
        self,
        board: Any,
        player: Any,
        *,
        top_n: Optional[int] = 3,
        include_all: bool = True,
        include_debug: bool = False,
        **overrides: Any,
    ) -> Dict[str, Any]:
        return rank_strategies_for_player(
            board,
            player,
            requirements=self.requirements,
            top_n=top_n,
            include_all=include_all,
            include_debug=include_debug,
            **self._timing_kwargs(overrides),
        )

    def rank_for_game(
        self,
        game: Any,
        *,
        top_n: int = 3,
        include_all: bool = False,
        include_debug: bool = False,
        player_ids: Optional[Iterable[int]] = None,
        **overrides: Any,
    ) -> Dict[str, Any]:
        return build_strategy_timing_report(
            game,
            requirements=self.requirements,
            top_n=top_n,
            include_all=include_all,
            include_debug=include_debug,
            player_ids=player_ids,
            **self._timing_kwargs(overrides),
        )


# ──────────────────────────────────────────────────────────────────────────────
# CSV export helpers
# ──────────────────────────────────────────────────────────────────────────────

def _csv_join(values: Any) -> str:
    if values is None:
        return ""
    if isinstance(values, str):
        return values.replace("Wool", "Sheep")
    try:
        iterator = iter(values)
    except TypeError:
        return str(values).replace("Wool", "Sheep")
    return "; ".join(str(v).replace("Wool", "Sheep") for v in iterator)


def _csv_vector(values: Any) -> str:
    if values is None:
        return ""
    if isinstance(values, (list, tuple)):
        out = []
        for value in values[:5]:
            try:
                f = float(value)
                out.append(str(int(round(f))) if abs(f - round(f)) < 1e-9 else str(round(f, 3)))
            except Exception:
                out.append(str(value))
        return "[" + ",".join(out) + "]"
    return str(values)


def _csv_vector_columns(prefix: str, values: Any) -> Dict[str, Any]:
    vec = list(values or [])[:5]
    if len(vec) < 5:
        vec.extend([""] * (5 - len(vec)))
    return {f"{prefix}_{RESOURCE_ORDER_NAMES[i].lower()}": vec[i] for i in range(5)}


def iter_strategy_timing_report_csv_rows(
    report: Mapping[str, Any],
    *,
    top_n: Optional[int] = None,
) -> Iterable[Dict[str, Any]]:
    """Yield one flat CSV row per player/top-strategy from a timing report."""
    by_player = report.get("by_player", {}) or {}

    for player_key in sorted(by_player, key=lambda k: safe_int(k, 9999)):
        player_report = by_player[player_key] or {}
        player_info = player_report.get("player", {}) or {}
        strategies = list(player_report.get("top_strategies", []) or [])
        if top_n is not None:
            strategies = strategies[: max(0, int(top_n))]

        for row in strategies:
            remaining = row.get("remaining", {}) or {}
            progress = remaining.get("progress", {}) or {}
            summary = row.get("strategy_summary", {}) or {}
            trades = row.get("player_trade_opportunities", []) or []
            trade_diag = row.get("trade_diagnostics", {}) or {}
            player_special = player_report.get("special_strategy_viability", {}) or {}
            lr_viability = player_special.get("longest_road", {}) or {}
            la_viability = player_special.get("largest_army", {}) or {}

            flat: Dict[str, Any] = {
                "round": report.get("round"),
                "turn": report.get("turn"),
                "phase": report.get("phase"),
                "player_id": player_info.get("player_id"),
                "color": player_info.get("color"),
                "current_hand": _csv_vector(player_info.get("current_hand")),
                "production_pips": _csv_vector(player_info.get("production_pips")),
                "trade_rates": _csv_vector(player_info.get("trade_rates")),
                "ports": _csv_join(player_info.get("ports")),
                "settlements": _csv_join(player_info.get("settlements")),
                "cities": _csv_join(player_info.get("cities")),
                "roads_count": player_info.get("roads_count"),
                "dev_card_progress": player_info.get("dev_card_progress"),
                "rank": row.get("rank"),
                "way_id": row.get("way_id"),
                "turns": row.get("turns"),
                "found": row.get("found"),
                "confidence": row.get("confidence"),
                "confidence_label": row.get("confidence_label"),
                "need_vector": _csv_vector(row.get("need_vector")),
                "total_cards": remaining.get("total_cards"),
                "expected_hand": _csv_vector(row.get("expected_hand")),
                "tags": _csv_join(row.get("tags")),
                "reason": _csv_join(row.get("reason")),
                "bottlenecks": _csv_join(row.get("bottlenecks")),
                "trade_dependency": _csv_join(row.get("trade_dependency")),
                "discrete_bank_trade_payable": trade_diag.get("discrete_bank_trade_payable"),
                "discrete_trade_gap": trade_diag.get("discrete_trade_gap"),
                "trade_margin": trade_diag.get("trade_margin"),
                "continuous_trade_margin": trade_diag.get("continuous_trade_margin"),
                "continuous_only_trade_estimate": trade_diag.get("continuous_only_trade_estimate"),
                "trade_warning": trade_diag.get("trade_warning"),
                "remaining_new_settlements": remaining.get("remaining_new_settlements"),
                "remaining_city_upgrades": remaining.get("remaining_city_upgrades"),
                "remaining_roads_to_build": remaining.get("remaining_roads_to_build"),
                "remaining_dev_cards_to_buy": remaining.get("remaining_dev_cards_to_buy"),
                "current_settlements": progress.get("current_settlements"),
                "current_cities": progress.get("current_cities"),
                "current_total_buildings": progress.get("current_total_buildings"),
                "target_cities": progress.get("target_cities"),
                "target_settlements": progress.get("target_settlements"),
                "target_total_buildings": progress.get("target_total_buildings"),
                "current_road_credit": progress.get("current_roads_credit", progress.get("current_road_credit")),
                "current_dev_card_credit": progress.get("current_dev_card_credit"),
                "road_progress_is_count_based": progress.get("road_progress_is_count_based"),
                "dev_card_progress_subtracted": progress.get("dev_card_progress_subtracted"),
                "warnings": _csv_join(remaining.get("warnings")),
                "longest_road": summary.get("longest_road"),
                "largest_army": summary.get("largest_army", summary.get("biggest_army", False)),
                "longest_road_viable": lr_viability.get("viable"),
                "longest_road_status": lr_viability.get("status"),
                "longest_road_reason": lr_viability.get("reason"),
                "largest_army_viable": la_viability.get("viable"),
                "largest_army_status": la_viability.get("status"),
                "largest_army_reason": la_viability.get("reason"),
                "cities_required": summary.get("cities"),
                "settlements_required": summary.get("settlements"),
                "vp_cards": summary.get("victory_point_cards"),
                "dev_cards_to_buy": summary.get("development_cards_to_buy"),
                "article_min_cost": summary.get("article_min_cost"),
                "total_vp": summary.get("total_victory_points"),
                "twelve_point_edge_case": summary.get("twelve_point_edge_case"),
                "player_trade_opportunities_count": len(trades),
                "player_trade_opportunities": _csv_join(
                    f"P{trade.get('counterparty_player_id')}: give {trade.get('give')} / receive {trade.get('receive')} ({trade.get('status')})"
                    for trade in trades
                ),
            }

            flat.update(_csv_vector_columns("hand", player_info.get("current_hand")))
            flat.update(_csv_vector_columns("pips", player_info.get("production_pips")))
            flat.update(_csv_vector_columns("need", row.get("need_vector")))
            flat.update(_csv_vector_columns("expected", row.get("expected_hand")))
            yield flat


def write_strategy_timing_report_csv(
    report: Mapping[str, Any] | os.PathLike[str] | str,
    csv_path: os.PathLike[str] | str,
    *,
    top_n: Optional[int] = None,
) -> Path:
    """Write a flat CSV summary from a StrategyTiming report dict or JSON file."""
    if isinstance(report, (str, os.PathLike)):
        with Path(report).open("r", encoding="utf-8") as f:
            report_data = json.load(f)
    else:
        report_data = report

    rows = list(iter_strategy_timing_report_csv_rows(report_data, top_n=top_n))
    output_path = Path(csv_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if not rows:
        output_path.write_text("", encoding="utf-8")
        return output_path

    with output_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)

    return output_path


# ──────────────────────────────────────────────────────────────────────────────
# Human-readable formatting helpers
# ──────────────────────────────────────────────────────────────────────────────

def format_player_top_strategies(report: Mapping[str, Any], *, limit: int = 3) -> str:
    """Return a compact text summary for logs/console/debug UI."""
    player = report.get("player", {}) or {}
    player_id = player.get("player_id", "?")
    color = player.get("color", "")
    top = list(report.get("top_strategies", []) or [])[:limit]

    lines = [f"Player {player_id} {color} — Top {len(top)} strategy timings"]
    for row in top:
        way_id = row.get("way_id")
        turns = row.get("turns")
        found = row.get("found")
        confidence_label = row.get("confidence_label", "")
        need = row.get("need_compact", "")
        tags = ", ".join(row.get("tags", []) or [])
        reason = row.get("reason", "")
        marker = "" if found else " (not found within max turns)"
        lines.append(
            f"#{row.get('rank')} Way {way_id}: {turns} turns{marker} | "
            f"confidence={confidence_label} | need={need} | {tags} | {reason}"
        )

    best_non_trade = report.get("best_non_trade_dependent_strategy")
    if best_non_trade:
        lines.append(
            "Best non-trade-dependent: "
            f"Way {best_non_trade.get('way_id')} at {best_non_trade.get('turns')} turns"
        )

    return "\n".join(lines)


def format_game_top_strategies(report: Mapping[str, Any], *, limit: int = 3) -> str:
    """Return a compact text summary for a multi-player timing report."""
    lines = [
        f"Strategy timing report — round={report.get('round')} turn={report.get('turn')} "
        f"phase={report.get('phase')}"
    ]
    by_player = report.get("by_player", {}) or {}
    for _, player_report in sorted(by_player.items(), key=lambda kv: safe_int(kv[0], 0)):
        lines.append("")
        lines.append(format_player_top_strategies(player_report, limit=limit))
    return "\n".join(lines)


__all__ = [
    "StrategyRequirement",
    "PlayerStrategyState",
    "SpecialStrategyViability",
    "StrategyRemainingNeed",
    "StrategyTimingRow",
    "StrategyTimingEngine",
    "build_player_strategy_state",
    "calculate_remaining_need",
    "compute_discrete_trade_diagnostics",
    "estimate_resource_requirement_time",
    "evaluate_largest_army_viability",
    "evaluate_longest_road_viability",
    "evaluate_special_strategy_viability",
    "format_game_top_strategies",
    "format_player_top_strategies",
    "iter_strategy_timing_report_csv_rows",
    "write_strategy_timing_report_csv",
    "expected_development_card_buys",
    "load_strategy_requirements",
    "rank_strategies_for_player",
    "rank_strategies_for_player_state",
    "build_strategy_timing_report",
    "resolve_strategy_requirements_path",
    "strategy_cost_from_components",
]
